import os
import shutil
import tabula
import pandas as pd
import openpyxl
from openpyxl import Workbook
import csv
import undetected_chromedriver
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

# TODO: add warning list for exceptions

"""definitions"""
User = "flyin"
ProjectPath = os.path.dirname(os.path.realpath(__file__))

BnbConsultUrl = 'https://cri.nbb.be/bc9/web/catalog?execution=e1s1'

OutputFolderPath   = ProjectPath + r'\Output'
TemplateExcelPath  = ProjectPath + r'\tempvlook_full.xlsx'
CopyhereFolderPath = ProjectPath + r'\Copy here'
FormExcelPath      = ProjectPath + r'\Bnbtemplateform.xlsx'   # Excel with company numbers

CompanyNumberList = []
YearstoScrape = 1

"""temporary testing definitions"""
#CompanyNumberList = ['0664.642.317', '0568.484.039']

"""replace temporary testing definitions by files in copy here folder - if any"""
if os.path.exists(CopyhereFolderPath):
    FileList = os.listdir(CopyhereFolderPath)
    for i in range(len(FileList)):
       if FileList[i].lower().endswith('.xlsx'):
            FormExcelPath = CopyhereFolderPath+'\\'+FileList[i]

def BnbScrape(Companynumber):
    print('start scrape')
    TemporaryFolderPath = OutputFolderPath + r'\temp'
    if not os.path.exists(OutputFolderPath):
        os.makedirs(OutputFolderPath)
    if not os.path.exists(TemporaryFolderPath):
        os.makedirs(TemporaryFolderPath)
    print('folders OK')
    profile = webdriver.FirefoxProfile()
    mime_types = "application/pdf,application/vnd.adobe.xfdf,application/vnd.fdf,application/vnd.adobe.xdp+xml"
    profile.set_preference("browser.download.folderList", 2)                        #not use default Downloads directory
    profile.set_preference("browser.helperApps.alwaysAsk.force", False)
    profile.set_preference("browser.download.manager.showWhenStarting", False)      #turns of showing download progress
    profile.set_preference("browser.download.manager.showAlertOnComplete", False)
    profile.set_preference("browser.download.dir", TemporaryFolderPath)                   #sets  directory for downloads
    profile.set_preference("browser.helperApps.neverAsk.saveToDisk", mime_types)
    profile.set_preference("plugin.disable_full_page_plugin_for_types", mime_types)
    profile.set_preference("pdfjs.disabled", True)
    BnbBrowser = webdriver.Firefox(firefox_profile=profile)
    BnbBrowser.implicitly_wait(10)
    BnbBrowser.get(BnbConsultUrl)

    NumberInput = BnbBrowser.find_element(By.CSS_SELECTOR, '#page_searchForm\:j_id3\:generated_number_2_component')
    NumberInput.send_keys(Companynumber)
    NumberInput.send_keys(Keys.ENTER)

    for t in range(YearstoScrape):
        try:
            DownloadButton = BnbBrowser.find_element(By.ID,'j_idt131:j_idt165:'+str(t)+':generated_pdfDownload_0_cell')
            DownloadButton.click()
            time.sleep(2)
        except:
            time.sleep(1)
    print('downloads launched')
    CompanyName = BnbBrowser.find_element(By.CSS_SELECTOR,'#j_idt131\:j_idt137 > tbody:nth-child(1) > tr:nth-child(2) > td:nth-child(2)')

    #wait for downloads to complete
    seconds = 0
    dl_wait = True
    while dl_wait and seconds < 60:
        time.sleep(1)
        for fname in os.listdir(TemporaryFolderPath):
            if fname.endswith('.part'):
                dl_wait = True
                seconds += 1
            else:
                dl_wait = False
                print('downloads done')

    os.rename(TemporaryFolderPath, OutputFolderPath + '\\' + str(CompanyName.text))
    print('rename OK')
    BnbBrowser.quit()
    print('quit OK')
    time.sleep(2)

def remove_empty(line):
    result = []
    for i in range(len(line)):
        if line[i] != "":
            result.append(line[i])
    return result

def ExtractPDF(Pdf,i, CompanyFolderPath, destinationwb):
    PDFPath = CompanyFolderPath+'\\'+str(Pdf)
    OutputCSVPath = CompanyFolderPath+'\\scrapeCSV'+str(i)+'.csv'

    tabula.convert_into(PDFPath, OutputCSVPath, pages='all', output_format="csv", stream=True)

    ws = destinationwb['Acc('+str(i)+')']
    with open(OutputCSVPath) as CSV:
        reader = csv.reader(CSV, delimiter=',')
        for row in reader:
            row = remove_empty(row)
            for i in range(len(row)):
                row[i] = row[i].replace(".","")
                try:
                    row[i] = float(row[i])
                except:
                    pass
            ws.append(row)

def ReadExcelInput(path):
    wb = openpyxl.load_workbook(path)
    FormSheet = wb['Form']
    global YearstoScrape
    YearstoScrape = FormSheet.cell(row=5,column=5).value
    for i in range(8,FormSheet.max_row):
        cell = FormSheet.cell(row=i,column=5)
        CompanyNr = cell.value
        if CompanyNr != None:
            if len(CompanyNr) > 2:
                      CompanyNumberList.append(CompanyNr)

def Execute():
    ReadExcelInput(FormExcelPath)
    print(CompanyNumberList)
    print('years to scrape : ' + str(YearstoScrape))

    print('deleting old outputs')
    for FolderName in os.listdir(OutputFolderPath):
        print(FolderName)
        FolderPath = OutputFolderPath + '\\' + FolderName
        shutil.rmtree(FolderPath)


    for j in range(len(CompanyNumberList)):
        BnbScrape(CompanyNumberList[j])

    CompanyList = os.listdir(OutputFolderPath)
    for Company in CompanyList:
        print('start extract '+str(Company))
        i = 1
        CompanyFolderPath = OutputFolderPath + '\\' + str(Company)
        PdfList = os.listdir(CompanyFolderPath)
        wb = openpyxl.load_workbook(TemplateExcelPath)
        for Pdf in PdfList:
            print('start extract ' + str(Pdf))
            try:
                ExtractPDF(Pdf, i, CompanyFolderPath, wb)
                i += 1
            except:
                print('extract '+str(Company) + ' ' + str(Pdf) + ' failed')
                i +=1
        OutputExcelPath = CompanyFolderPath + '\\'+str(Company)+'ScrapedAccounts.xlsx'
        print('end extract'+str(Company))
        wb.save(OutputExcelPath)

if __name__ == '__main__':
    Execute()